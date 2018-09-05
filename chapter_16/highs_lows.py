import csv
import openpyxl
import re
from datetime import datetime

from matplotlib import pyplot as plt


# Get dates, high, and low temperatures from file.
def main():
    filename = 'death_valley_2014.csv'
    with open(filename) as f:
        reader = csv.reader(f)
        header_row = next(reader)

        dates, highs, lows = [], [], []
        for row in reader:
            try:
                current_date = datetime.strptime(row[0], "%Y-%m-%d")
                high = int(row[1])
                low = int(row[3])
            except ValueError:
                print(current_date, 'missing data')
            else:
                dates.append(current_date)
                highs.append(high)
                lows.append(low)

    # Plot data.
    fig = plt.figure(dpi=128, figsize=(10, 6))
    plt.plot(dates, highs, c='red', alpha=0.5)
    plt.plot(dates, lows, c='blue', alpha=0.5)
    plt.fill_between(dates, highs, lows, facecolor='blue', alpha=0.1)

    # Format plot.
    title = "Daily high and low temperatures - 2014\nDeath Valley, CA"
    plt.title(title, fontsize=20)
    plt.xlabel('', fontsize=16)
    fig.autofmt_xdate()
    plt.ylabel("Temperature (F)", fontsize=16)
    plt.tick_params(axis='both', which='major', labelsize=16)

    cvs2xls(filename)
    plt.show()


def cvs2xls(filename):
    patternname = "".join(re.findall(r'(.+?)\.', filename, flags=re.IGNORECASE))
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = patternname
    with open(filename) as f:
        reader = csv.reader(f)
        header_row = next(reader)
        i = 1
        for j, y in enumerate(header_row):
            ws[chr(65 + j) + str(i)] = y
        for row in (reader):
            i += 1
            for j, y in enumerate(row):
                ws[chr(65 + j) + str(i)] = y

    wb.save(patternname + '.xlsx')


if __name__ == '__main__':
    main()
